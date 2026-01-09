#!/usr/bin/env python3
"""
テンプレート管理モジュール
Excelファイルを解析してテンプレート設定を自動生成
"""
import json
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def analyze_excel_template(excel_path: Path) -> dict:
    """
    Excelファイルを解析してテンプレート構造を抽出
    
    Returns:
        {
            "name": "テンプレート名",
            "sheet": "シート名",
            "structure": [
                {"cell": "A1", "value": "タイトル", "is_header": True, "row": 1, "col": 1},
                {"cell": "B2", "value": "", "is_header": False, "row": 2, "col": 2},
                ...
            ],
            "suggested_sections": [
                {"key": "section_1", "label": "項目名", "cell": "B2"},
                ...
            ]
        }
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    sheet_name = ws.title
    
    structure = []
    suggested_sections = []
    
    # 全セルをスキャン
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell_ref = f"{get_column_letter(col)}{row}"
            value = cell.value
            
            # セル情報を記録
            cell_info = {
                "cell": cell_ref,
                "value": str(value) if value else "",
                "row": row,
                "col": col,
                "is_merged": False,
                "has_content": value is not None and str(value).strip() != "",
                "font_bold": cell.font.bold if cell.font else False,
                "fill_color": cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
            }
            
            structure.append(cell_info)
            
            # 入力フィールドを推測（空セルで、左または上にラベルがあるもの）
            if not cell_info["has_content"]:
                # 左のセルをチェック
                if col > 1:
                    left_cell = ws.cell(row=row, column=col-1)
                    if left_cell.value and str(left_cell.value).strip():
                        label = str(left_cell.value).strip()
                        # 【】や「」で囲まれたラベルを抽出
                        label_clean = re.sub(r'[【】「」\[\]（）()]', '', label)
                        key = _generate_key(label_clean)
                        suggested_sections.append({
                            "key": key,
                            "label": label_clean,
                            "cell": cell_ref,
                            "row": row,
                            "col": col
                        })
    
    # マージされたセルを確認
    for merged_range in ws.merged_cells.ranges:
        for cell_info in structure:
            if cell_info["cell"] in [str(c) for c in merged_range.cells]:
                cell_info["is_merged"] = True
    
    # テンプレート名を推測（最初の非空セル）
    template_name = "新規テンプレート"
    for cell_info in structure:
        if cell_info["has_content"] and cell_info["row"] == 1:
            template_name = cell_info["value"][:30]  # 最大30文字
            break
    
    return {
        "name": template_name,
        "sheet": sheet_name,
        "excel_file": excel_path.name,
        "structure": structure,
        "suggested_sections": suggested_sections,
        "max_row": ws.max_row,
        "max_col": ws.max_column
    }


def _generate_key(label: str) -> str:
    """ラベルからキーを生成"""
    # 日本語をローマ字に変換（簡易版）
    # 実際はより高度な変換が必要
    import unicodedata
    
    # 基本的なクリーンアップ
    key = label.lower().strip()
    key = re.sub(r'\s+', '_', key)
    key = re.sub(r'[^\w]', '', key)
    
    # 日本語が含まれる場合はそのまま使用（または番号を付ける）
    if not key or not key.isascii():
        key = f"field_{hash(label) % 10000:04d}"
    
    return key


def create_template_config(
    excel_path: Path,
    template_name: str,
    sections: list,
    style_rules: list = None
) -> dict:
    """
    テンプレート設定ファイル（JSON）を生成
    
    Args:
        excel_path: Excelファイルのパス
        template_name: テンプレート名
        sections: セクション設定のリスト
            [{"key": "datetime", "label": "日時", "cell": "B2"}, ...]
        style_rules: スタイルルール（オプション）
    
    Returns:
        テンプレート設定の辞書
    """
    if style_rules is None:
        style_rules = [
            "重複や雑談は除外",
            "箇条書き",
            "決定事項・課題・次アクションがあれば必ず含める",
            "不明点は『（要確認）』と付ける"
        ]
    
    config = {
        "name": template_name,
        "excel_template": f"excel_templates/{excel_path.name}",
        "sheet": load_workbook(excel_path).active.title,
        "sections": sections,
        "chatgpt_prompt": {
            "style_rules": style_rules
        }
    }
    
    return config


def save_template_config(config: dict, output_dir: Path) -> Path:
    """テンプレート設定をJSONファイルとして保存"""
    # ファイル名を生成
    filename = re.sub(r'[^\w\-]', '_', config["name"]) + ".json"
    output_path = output_dir / filename
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    return output_path


def preview_template_mapping(excel_path: Path, sections: list) -> dict:
    """
    テンプレートマッピングのプレビューを生成
    
    Returns:
        各セクションの位置とサンプルデータを含む辞書
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    
    preview = {
        "sections": [],
        "preview_data": {}
    }
    
    for section in sections:
        cell_ref = section["cell"]
        cell = ws[cell_ref]
        
        # セルの周辺情報を取得
        row, col = cell.row, cell.column
        
        section_preview = {
            "key": section["key"],
            "label": section["label"],
            "cell": cell_ref,
            "current_value": str(cell.value) if cell.value else "(空)",
            "row_height": ws.row_dimensions[row].height,
            "col_width": ws.column_dimensions[get_column_letter(col)].width
        }
        
        preview["sections"].append(section_preview)
        preview["preview_data"][section["key"]] = f"【{section['label']}の内容がここに入ります】"
    
    return preview


# CLI用の関数
def interactive_template_setup(excel_path: str) -> dict:
    """
    対話形式でテンプレートを設定（CLI用）
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {excel_path}")
    
    print(f"\n📊 Excelファイルを解析中: {excel_path.name}")
    analysis = analyze_excel_template(excel_path)
    
    print(f"\n📋 テンプレート名: {analysis['name']}")
    print(f"📄 シート名: {analysis['sheet']}")
    print(f"📐 サイズ: {analysis['max_row']}行 x {analysis['max_col']}列")
    
    print("\n🔍 検出された入力候補フィールド:")
    for i, section in enumerate(analysis["suggested_sections"], 1):
        print(f"  {i}. [{section['cell']}] {section['label']}")
    
    return analysis


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("使用法: python template_manager.py <Excelファイルパス>")
        sys.exit(1)
    
    result = interactive_template_setup(sys.argv[1])
    print("\n📝 解析結果（JSON）:")
    print(json.dumps(result, ensure_ascii=False, indent=2))
