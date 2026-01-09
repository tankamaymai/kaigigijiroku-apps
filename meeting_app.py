#!/usr/bin/env python3
"""
議事録自動生成アプリ（ワンクリック版）
- Whisperで文字起こし
- ChatGPT APIで自動要約
- Excel自動出力
"""
import os
import json
import glob
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from datetime import datetime
import threading

# OpenAI API
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

# アプリケーションのディレクトリを取得
APP_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(APP_DIR, "templates")
OUTPUT_DIR = os.path.join(APP_DIR, "output")
CONFIG_FILE = os.path.join(APP_DIR, "config.json")

# outputフォルダがなければ作成
os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_config() -> dict:
    """設定ファイルを読み込む"""
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"openai_api_key": "", "whisper_model": "medium", "gpt_model": "gpt-4o-mini"}


def save_config(config: dict):
    """設定ファイルを保存"""
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def list_templates():
    """テンプレートJSONファイルの一覧を取得"""
    files = sorted(glob.glob(os.path.join(TEMPLATE_DIR, "*.json")))
    return [os.path.basename(f) for f in files]


def load_template(filename: str) -> dict:
    """テンプレートJSONを読み込む"""
    path = os.path.join(TEMPLATE_DIR, filename)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def run_whisper(audio_path: str, model_name: str = "medium") -> str:
    """
    whisper CLI を呼び出して、文字起こしを行う
    戻り値：生成された .txt のパス
    """
    subprocess.run(
        ["whisper", audio_path, "--language", "Japanese", "--model", model_name, "--output_dir", OUTPUT_DIR],
        check=True
    )

    base = os.path.splitext(os.path.basename(audio_path))[0]
    txt_path = os.path.join(OUTPUT_DIR, f"{base}.txt")
    if not os.path.exists(txt_path):
        raise FileNotFoundError(f"文字起こしファイルが見つかりません: {txt_path}")
    return txt_path


def build_prompt(tpl: dict, transcript: str) -> str:
    """ChatGPT用のプロンプトを構築"""
    sections = "\n".join([f"- {s['label']}（key: {s['key']}）" for s in tpl["sections"]])
    rules = "\n".join([f"- {r}" for r in tpl.get("chatgpt_prompt", {}).get("style_rules", [])])

    return f"""以下の会議文字起こしを、指定のテンプレート構造に沿って整理してください。

【テンプレート名】
{tpl["name"]}

【出力する項目】
{sections}

【ルール】
{rules}

【出力形式（厳守）】
必ずJSONだけを返してください。キーは sections の key を使ってください。
コードブロック（```）は使わず、純粋なJSONのみを返してください。

例：
{{
  "datetime": "2026年1月8日",
  "shinkomonji": "・○○の報告\\n・△△について協議",
  "fukuokawajiro": "・□□の進捗報告",
  "shinyukuhashi": "・◇◇の改善提案"
}}

【文字起こし】
{transcript}
"""


def call_chatgpt(api_key: str, prompt: str, model: str = "gpt-4o-mini") -> str:
    """ChatGPT APIを呼び出して要約を取得"""
    client = OpenAI(api_key=api_key)
    
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "あなたは議事録作成のプロフェッショナルです。会議の内容を正確かつ簡潔にまとめてください。"},
            {"role": "user", "content": prompt}
        ],
        temperature=0.3,
        max_tokens=4000
    )
    
    return response.choices[0].message.content


def parse_json_response(response: str) -> dict:
    """ChatGPTのレスポンスからJSONをパース"""
    # コードブロックを除去
    if "```json" in response:
        response = response.split("```json")[1]
    if "```" in response:
        response = response.split("```")[0]
    response = response.strip()
    
    return json.loads(response)


def write_excel(tpl: dict, data: dict, out_name: str) -> str:
    """テンプレートExcelにデータを書き込んで保存"""
    xlsx_path = os.path.join(APP_DIR, tpl["excel_template"])
    wb = load_workbook(xlsx_path)
    ws = wb[tpl["sheet"]]

    for s in tpl["sections"]:
        key = s["key"]
        cell = s["cell"]
        ws[cell] = data.get(key, "")

    out_path = os.path.join(OUTPUT_DIR, out_name)
    wb.save(out_path)
    return out_path


# ================ UI ================

class MeetingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📝 議事録メーカー（ワンクリック自動生成）")
        self.root.geometry("950x750")
        self.root.configure(bg="#1a1a2e")
        
        # 設定読み込み
        self.config = load_config()
        
        # 変数
        self.templates = list_templates()
        if not self.templates:
            messagebox.showerror("エラー", "templates/ に json がありません。")
            raise SystemExit(1)
        
        self.template_var = tk.StringVar(value=self.templates[0])
        self.audio_path_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="準備完了")
        self.progress_var = tk.StringVar(value="")
        self.api_key_var = tk.StringVar(value=self.config.get("openai_api_key", ""))
        self.whisper_model_var = tk.StringVar(value=self.config.get("whisper_model", "medium"))
        self.gpt_model_var = tk.StringVar(value=self.config.get("gpt_model", "gpt-4o-mini"))
        
        self.setup_ui()
    
    def setup_ui(self):
        """UIをセットアップ"""
        # カスタムスタイル
        style = ttk.Style()
        style.theme_use('clam')
        
        # カラーパレット
        bg_color = "#1a1a2e"
        card_color = "#16213e"
        accent_color = "#e94560"
        text_color = "#eaeaea"
        
        style.configure("TFrame", background=bg_color)
        style.configure("Card.TFrame", background=card_color)
        style.configure("TLabel", background=bg_color, foreground=text_color, font=("Helvetica", 11))
        style.configure("Card.TLabel", background=card_color, foreground=text_color)
        style.configure("Header.TLabel", background=bg_color, foreground="#ffffff", font=("Helvetica", 18, "bold"))
        style.configure("Status.TLabel", background=bg_color, foreground="#888888", font=("Helvetica", 10))
        style.configure("TButton", font=("Helvetica", 11), padding=10)
        style.configure("Accent.TButton", font=("Helvetica", 13, "bold"), padding=15)
        style.configure("TLabelframe", background=card_color, foreground=text_color)
        style.configure("TLabelframe.Label", background=card_color, foreground=accent_color, font=("Helvetica", 11, "bold"))
        style.configure("TCombobox", font=("Helvetica", 11))
        style.configure("TEntry", font=("Helvetica", 11))
        
        # メインフレーム
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # タイトル
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = ttk.Label(title_frame, text="📝 議事録メーカー", style="Header.TLabel")
        title_label.pack()
        
        subtitle_label = ttk.Label(title_frame, text="音声ファイルを選んでワンクリックで議事録完成", 
                                   style="Status.TLabel")
        subtitle_label.pack()
        
        # 設定セクション
        settings_frame = ttk.LabelFrame(main_frame, text="⚙️ 設定", padding="15")
        settings_frame.pack(fill=tk.X, pady=10)
        
        # API Key
        api_frame = ttk.Frame(settings_frame, style="Card.TFrame")
        api_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(api_frame, text="OpenAI API Key:", style="Card.TLabel").pack(side=tk.LEFT)
        api_entry = ttk.Entry(api_frame, textvariable=self.api_key_var, width=50, show="*")
        api_entry.pack(side=tk.LEFT, padx=10)
        
        ttk.Button(api_frame, text="保存", command=self.save_api_key).pack(side=tk.LEFT)
        
        # モデル設定
        model_frame = ttk.Frame(settings_frame, style="Card.TFrame")
        model_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(model_frame, text="Whisperモデル:", style="Card.TLabel").pack(side=tk.LEFT)
        whisper_combo = ttk.Combobox(model_frame, textvariable=self.whisper_model_var,
                                     values=["tiny", "base", "small", "medium", "large"],
                                     state="readonly", width=10)
        whisper_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(model_frame, text="  GPTモデル:", style="Card.TLabel").pack(side=tk.LEFT)
        gpt_combo = ttk.Combobox(model_frame, textvariable=self.gpt_model_var,
                                 values=["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"],
                                 state="readonly", width=15)
        gpt_combo.pack(side=tk.LEFT, padx=5)
        
        # テンプレート選択
        template_frame = ttk.LabelFrame(main_frame, text="📋 テンプレート選択", padding="15")
        template_frame.pack(fill=tk.X, pady=10)
        
        template_combo = ttk.Combobox(template_frame, textvariable=self.template_var,
                                      values=self.templates, state="readonly", width=50)
        template_combo.pack(side=tk.LEFT)
        
        # 音声ファイル選択
        audio_frame = ttk.LabelFrame(main_frame, text="🎤 音声ファイル", padding="15")
        audio_frame.pack(fill=tk.X, pady=10)
        
        audio_btn = ttk.Button(audio_frame, text="📁 音声ファイルを選ぶ", command=self.choose_audio)
        audio_btn.pack(side=tk.LEFT)
        
        audio_label = ttk.Label(audio_frame, textvariable=self.audio_path_var, 
                                style="Card.TLabel", wraplength=600)
        audio_label.pack(side=tk.LEFT, padx=15)
        
        # ワンクリック実行ボタン
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=20)
        
        self.run_button = ttk.Button(action_frame, text="🚀 ワンクリックで議事録作成", 
                                     command=self.run_one_click, style="Accent.TButton")
        self.run_button.pack(expand=True)
        
        # プログレス表示
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=10)
        
        self.progress_label = ttk.Label(progress_frame, textvariable=self.progress_var,
                                        style="Status.TLabel", font=("Helvetica", 12))
        self.progress_label.pack()
        
        # ログ表示エリア
        log_frame = ttk.LabelFrame(main_frame, text="📜 処理ログ", padding="15")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.log_text = tk.Text(log_frame, height=12, wrap=tk.WORD, 
                                font=("Monaco", 10), bg="#0f0f23", fg="#00ff00",
                                insertbackground="#00ff00")
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # スクロールバー
        scrollbar = ttk.Scrollbar(self.log_text, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # フッター
        footer_frame = ttk.Frame(main_frame)
        footer_frame.pack(fill=tk.X, pady=(10, 0))
        
        open_folder_btn = ttk.Button(footer_frame, text="📂 出力フォルダを開く", 
                                     command=self.open_output_folder)
        open_folder_btn.pack(side=tk.LEFT)
        
        status_label = ttk.Label(footer_frame, textvariable=self.status_var, style="Status.TLabel")
        status_label.pack(side=tk.RIGHT)
    
    def log(self, message: str):
        """ログを追加"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def save_api_key(self):
        """API Keyを保存"""
        self.config["openai_api_key"] = self.api_key_var.get()
        self.config["whisper_model"] = self.whisper_model_var.get()
        self.config["gpt_model"] = self.gpt_model_var.get()
        save_config(self.config)
        messagebox.showinfo("保存完了", "設定を保存しました。")
    
    def choose_audio(self):
        """音声ファイルを選択"""
        path = filedialog.askopenfilename(
            title="音声ファイルを選択",
            filetypes=[
                ("音声ファイル", "*.m4a *.mp3 *.wav *.mp4 *.webm *.ogg *.flac"),
                ("すべてのファイル", "*.*")
            ]
        )
        if path:
            self.audio_path_var.set(path)
            self.log(f"音声ファイル選択: {os.path.basename(path)}")
    
    def run_one_click(self):
        """ワンクリックで議事録作成"""
        # バリデーション
        if not self.audio_path_var.get():
            messagebox.showwarning("注意", "音声ファイルを選択してください。")
            return
        
        if not self.api_key_var.get():
            messagebox.showwarning("注意", "OpenAI API Keyを設定してください。")
            return
        
        if not OPENAI_AVAILABLE:
            messagebox.showerror("エラー", 
                "OpenAIライブラリがインストールされていません。\n\n"
                "以下を実行してください:\n"
                "pip install openai")
            return
        
        # 設定を保存
        self.save_api_key()
        
        # ボタンを無効化
        self.run_button.config(state=tk.DISABLED)
        
        # 別スレッドで実行
        thread = threading.Thread(target=self._run_pipeline)
        thread.start()
    
    def _run_pipeline(self):
        """パイプライン実行（別スレッド）"""
        try:
            self.log_text.delete("1.0", tk.END)
            
            # ステップ1: 文字起こし
            self.progress_var.set("🎙️ Step 1/3: 音声を文字起こし中...")
            self.log("=" * 50)
            self.log("🎙️ Whisperで文字起こしを開始...")
            self.log(f"   モデル: {self.whisper_model_var.get()}")
            self.log(f"   ファイル: {os.path.basename(self.audio_path_var.get())}")
            
            txt_path = run_whisper(self.audio_path_var.get(), self.whisper_model_var.get())
            
            with open(txt_path, "r", encoding="utf-8") as f:
                transcript = f.read().strip()
            
            self.log(f"✅ 文字起こし完了！ ({len(transcript)}文字)")
            self.log("")
            
            # ステップ2: ChatGPTで要約
            self.progress_var.set("🤖 Step 2/3: ChatGPTで議事録を作成中...")
            self.log("🤖 ChatGPT APIで議事録を作成中...")
            self.log(f"   モデル: {self.gpt_model_var.get()}")
            
            tpl = load_template(self.template_var.get())
            prompt = build_prompt(tpl, transcript)
            
            response = call_chatgpt(
                self.api_key_var.get(),
                prompt,
                self.gpt_model_var.get()
            )
            
            self.log("✅ ChatGPT応答を受信！")
            self.log("")
            
            # JSONをパース
            data = parse_json_response(response)
            self.log(f"✅ JSONパース成功！ (項目数: {len(data)})")
            
            # ステップ3: Excel生成
            self.progress_var.set("📊 Step 3/3: Excelファイルを生成中...")
            self.log("")
            self.log("📊 Excelファイルを生成中...")
            
            base = os.path.splitext(os.path.basename(self.audio_path_var.get()))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_name = f"{tpl['name']}_{base}_{timestamp}.xlsx"
            
            out_path = write_excel(tpl, data, out_name)
            
            self.log(f"✅ Excel出力完了！")
            self.log(f"   📁 {out_path}")
            self.log("")
            self.log("=" * 50)
            self.log("🎉 議事録作成が完了しました！")
            
            self.progress_var.set("✅ 完了！")
            self.status_var.set(f"出力: {out_name}")
            
            # 完了ダイアログ
            self.root.after(0, lambda: messagebox.showinfo(
                "🎉 完了！", 
                f"議事録を作成しました！\n\n📁 {out_path}"
            ))
            
        except subprocess.CalledProcessError:
            self.log("❌ エラー: Whisper実行失敗")
            self.progress_var.set("❌ エラー")
            self.root.after(0, lambda: messagebox.showerror("エラー",
                "Whisper実行に失敗しました。\n\n"
                "以下を確認してください:\n"
                "1. whisperがインストールされているか\n"
                "   pip install openai-whisper\n"
                "2. ffmpegがインストールされているか\n"
                "   brew install ffmpeg"))
        
        except json.JSONDecodeError as e:
            self.log(f"❌ エラー: JSONパース失敗 - {e}")
            self.progress_var.set("❌ エラー")
            self.root.after(0, lambda: messagebox.showerror("エラー",
                f"ChatGPTの応答をJSONとして解析できませんでした。\n\n"
                f"再度実行してください。"))
        
        except Exception as e:
            self.log(f"❌ エラー: {e}")
            self.progress_var.set("❌ エラー")
            self.root.after(0, lambda: messagebox.showerror("エラー", str(e)))
        
        finally:
            self.root.after(0, lambda: self.run_button.config(state=tk.NORMAL))
    
    def open_output_folder(self):
        """outputフォルダを開く"""
        subprocess.run(["open", OUTPUT_DIR])


def main():
    root = tk.Tk()
    app = MeetingApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
