"""
Excelテンプレートファイルを作成するスクリプト
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ワークブック作成
wb = Workbook()
ws = wb.active
ws.title = "原本"

# スタイル定義
title_font = Font(name="游ゴシック", size=14, bold=True)
header_font = Font(name="游ゴシック", size=12, bold=True)
normal_font = Font(name="游ゴシック", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

# 列幅設定
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 20

# タイトル行
ws['A1'] = "九州病院報告資料"
ws['A1'].font = title_font

# 日時行
ws['A2'] = "日時"
ws['A2'].font = header_font
ws['A2'].fill = header_fill
ws['A2'].border = thin_border
ws['B2'] = ""  # 日時が入る
ws['B2'].border = thin_border

# 各施設報告見出し
ws['A3'] = "各施設 報告（九州）"
ws['A3'].font = header_font
ws['A3'].fill = header_fill
ws['A3'].border = thin_border
ws.merge_cells('A3:C3')

# 施設ごとのセクション
facilities = [
    ("A4", "B4", "【新小文字】"),
    ("A5", "B5", "【福岡和白】"),
    ("A6", "B6", "【新行橋】"),
]

for label_cell, content_cell, facility_name in facilities:
    ws[label_cell] = facility_name
    ws[label_cell].font = header_font
    ws[label_cell].fill = header_fill
    ws[label_cell].border = thin_border
    ws[label_cell].alignment = Alignment(vertical='top')
    
    ws[content_cell] = ""  # 報告内容が入る
    ws[content_cell].font = normal_font
    ws[content_cell].border = thin_border
    ws[content_cell].alignment = Alignment(wrap_text=True, vertical='top')

# 行の高さを調整（報告内容が入る行は高くする）
ws.row_dimensions[4].height = 80
ws.row_dimensions[5].height = 80
ws.row_dimensions[6].height = 80

# 保存
output_path = "excel_templates/所属長会議まとめ.xlsx"
wb.save(output_path)
print(f"Excelテンプレートを作成しました: {output_path}")
